"""
Admin Routes

Admin-only endpoints for timesheet management.
Support users have limited access to approve trainee timesheets only (REQ-041).
"""

from datetime import datetime
from io import BytesIO, StringIO
import csv
from flask import Blueprint, request, session, send_file, current_app, Response
from ..models import (
    Timesheet,
    TimesheetEntry,
    Attachment,
    User,
    Note,
    TimesheetStatus,
    UserRole,
    PayPeriod,
)
from ..extensions import db
from ..utils.decorators import login_required, admin_required, can_approve
from ..utils.pay_periods import get_confirmed_pay_period

admin_bp = Blueprint("admin", __name__)


@admin_bp.route("/timesheets", methods=["GET"])
@login_required
@can_approve
def list_timesheets():
    """
    List submitted timesheets for approval.

    - Admin: sees all non-draft timesheets
    - Support: sees only trainee timesheets (REQ-041)

    Drafts (NEW status) are NOT visible.

    Query params:
        status: Filter by status (SUBMITTED, APPROVED, NEEDS_APPROVAL)
        user_id: Filter by user
        week_start: Filter by week (ISO date)
        page: Page number (default 1)
        per_page: Items per page (default 20)

    Returns:
        dict: Paginated list of timesheets with user info and view_mode
    """
    current_role = session.get("user", {}).get("role", "staff")
    is_support_only = (current_role == "support")
    
    # Base query - exclude drafts
    query = Timesheet.query.filter(Timesheet.status != TimesheetStatus.NEW)
    
    # REQ-041: Support users can only see trainee timesheets
    if is_support_only:
        # Join with User and filter to only trainee submitters
        query = query.join(User, Timesheet.user_id == User.id).filter(
            User.role == UserRole.TRAINEE
        )
    else:
        # Admin sees all - join with user for display name
        query = query.join(User, Timesheet.user_id == User.id)

    # Filter by status
    status = request.args.get("status")
    if status and status in [
        TimesheetStatus.SUBMITTED,
        TimesheetStatus.APPROVED,
        TimesheetStatus.NEEDS_APPROVAL,
    ]:
        query = query.filter(Timesheet.status == status)

    # Filter by user
    user_id = request.args.get("user_id")
    if user_id:
        query = query.filter(Timesheet.user_id == user_id)

    # Filter by week
    week_start = request.args.get("week_start")
    if week_start:
        query = query.filter(Timesheet.week_start == datetime.fromisoformat(week_start).date())

    # Filter by hour type (REQ-018)
    from ..models import TimesheetEntry
    hour_type = request.args.get("hour_type")
    if hour_type:
        if hour_type == "has_field":
            # Special case: show timesheets that have any Field hours
            query = query.filter(
                Timesheet.id.in_(
                    db.session.query(TimesheetEntry.timesheet_id)
                    .filter(TimesheetEntry.hour_type == "Field")
                    .distinct()
                )
            )
        else:
            # Filter by specific hour type
            query = query.filter(
                Timesheet.id.in_(
                    db.session.query(TimesheetEntry.timesheet_id)
                    .filter(TimesheetEntry.hour_type == hour_type)
                    .distinct()
                )
            )

    # Order by submitted_at (newest first)
    query = query.order_by(Timesheet.submitted_at.desc())

    # Paginate
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    pagination = query.paginate(page=page, per_page=per_page)

    timesheets = []
    for t in pagination.items:
        data = t.to_dict(include_entries=False)
        data["user"] = t.user.to_dict() if t.user else None
        timesheets.append(data)

    return {
        "timesheets": timesheets,
        "total": pagination.total,
        "page": page,
        "per_page": per_page,
        "pages": pagination.pages,
        # REQ-041: Tell frontend which view mode we're in
        "view_mode": "trainee_approvals" if is_support_only else "admin",
    }


@admin_bp.route("/reports/timesheet-data", methods=["GET"])
@login_required
@admin_required
def timesheet_data_report():
    """
    Raw data report view for timesheet entries (REQ-039).

    Query params:
        status: Filter by timesheet status
        user_id: Filter by user
        week_start: Filter by timesheet week (ISO date)
        hour_type: Filter by entry hour type
        start_date: Filter by entry date >= (ISO date)
        end_date: Filter by entry date <= (ISO date)
        page: Page number (default 1)
        per_page: Items per page (default 200)
    """
    query = TimesheetEntry.query.join(Timesheet).join(
        User, Timesheet.user_id == User.id
    )

    # Exclude drafts by default
    query = query.filter(Timesheet.status != TimesheetStatus.NEW)

    status = request.args.get("status")
    if status and status in [
        TimesheetStatus.SUBMITTED,
        TimesheetStatus.APPROVED,
        TimesheetStatus.NEEDS_APPROVAL,
    ]:
        query = query.filter(Timesheet.status == status)

    user_id = request.args.get("user_id")
    if user_id:
        query = query.filter(Timesheet.user_id == user_id)

    week_start = request.args.get("week_start")
    if week_start:
        query = query.filter(Timesheet.week_start == datetime.fromisoformat(week_start).date())

    hour_type = request.args.get("hour_type")
    if hour_type:
        query = query.filter(TimesheetEntry.hour_type == hour_type)

    start_date = request.args.get("start_date")
    if start_date:
        query = query.filter(
            TimesheetEntry.entry_date >= datetime.fromisoformat(start_date).date()
        )

    end_date = request.args.get("end_date")
    if end_date:
        query = query.filter(
            TimesheetEntry.entry_date <= datetime.fromisoformat(end_date).date()
        )

    query = query.order_by(
        TimesheetEntry.entry_date.desc(), TimesheetEntry.created_at.desc()
    )

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 200, type=int)
    pagination = query.paginate(page=page, per_page=per_page)

    rows = []
    for entry in pagination.items:
        timesheet = entry.timesheet
        user = timesheet.user if timesheet else None
        totals = timesheet.calculate_totals() if timesheet else {}
        reimbursement = "No"
        if timesheet and timesheet.reimbursement_needed:
            amount = float(timesheet.reimbursement_amount or 0)
            label = timesheet.reimbursement_type or "Reimbursement"
            reimbursement = f"{label}: ${amount:.2f}"

        rows.append(
            {
                "entry_id": entry.id,
                "timesheet_id": entry.timesheet_id,
                "employee": user.display_name if user else "Unknown",
                "email": user.email if user else "",
                "week_start": timesheet.week_start.isoformat()
                if timesheet
                else None,
                "entry_date": entry.entry_date.isoformat(),
                "hour_type": entry.hour_type,
                "hours": float(entry.hours),
                "status": timesheet.status if timesheet else None,
                "total_hours": float(totals.get("total", 0)),
                "payable_hours": float(totals.get("payable", 0)),
                "billable_hours": float(totals.get("billable", 0)),
                "unpaid_hours": float(totals.get("unpaid", 0)),
                "traveled": bool(timesheet.traveled) if timesheet else False,
                "expenses": bool(timesheet.has_expenses) if timesheet else False,
                "reimbursement": reimbursement,
                "attachments": timesheet.attachments.count() if timesheet else 0,
                "timesheet_created_at": timesheet.created_at.isoformat()
                if timesheet
                else None,
                "entry_created_at": entry.created_at.isoformat(),
            }
        )

    return {
        "rows": rows,
        "total": pagination.total,
        "page": page,
        "per_page": per_page,
        "pages": pagination.pages,
    }


def _can_access_timesheet(timesheet):
    """
    Check if current user can access a specific timesheet.
    
    REQ-041: Support can only access trainee timesheets.
    Admin can access all timesheets.
    
    Returns:
        tuple: (can_access: bool, error_response: tuple or None)
    """
    current_role = session.get("user", {}).get("role", "staff")
    
    # Admin can access everything
    if current_role == "admin":
        return True, None
    
    # Support can only access trainee timesheets
    if current_role == "support":
        if timesheet.user and timesheet.user.role == UserRole.TRAINEE:
            return True, None
        return False, ({"error": "You can only access trainee timesheets"}, 403)
    
    # Other roles shouldn't reach here due to @can_approve, but just in case
    return False, ({"error": "Access denied"}, 403)


@admin_bp.route("/timesheets/<timesheet_id>", methods=["GET"])
@login_required
@can_approve
def get_timesheet(timesheet_id):
    """
    Get a specific timesheet for approval review.
    
    REQ-041: Support users can only view trainee timesheets.

    Returns:
        dict: Timesheet with entries, attachments, and user info
    """
    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404

    # Cannot view drafts
    if timesheet.status == TimesheetStatus.NEW:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    data = timesheet.to_dict()
    period = get_confirmed_pay_period(timesheet.week_start)
    data["pay_period_confirmed"] = period is not None
    data["pay_period_confirmed_at"] = (
        period.confirmed_at.isoformat() if period else None
    )
    data["user"] = timesheet.user.to_dict() if timesheet.user else None
    data["notes"] = [n.to_dict() for n in timesheet.notes]

    return data


@admin_bp.route("/timesheets/<timesheet_id>/approve", methods=["POST"])
@login_required
@can_approve
def approve_timesheet(timesheet_id):
    """
    Approve a submitted timesheet.
    
    REQ-041: Support users can only approve trainee timesheets.

    Returns:
        dict: Updated timesheet
    """
    approver_id = session["user"]["id"]

    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    if timesheet.status not in [
        TimesheetStatus.SUBMITTED,
        TimesheetStatus.NEEDS_APPROVAL,
    ]:
        return {"error": "Timesheet cannot be approved from current status"}, 400

    if get_confirmed_pay_period(timesheet.week_start):
        return {"error": "Pay period has been confirmed and is locked"}, 400

    timesheet.status = TimesheetStatus.APPROVED
    timesheet.approved_at = datetime.utcnow()
    timesheet.approved_by = approver_id
    db.session.commit()

    # Send SMS notification
    from ..services.notification import NotificationService

    NotificationService.notify_approved(timesheet)

    return timesheet.to_dict()


@admin_bp.route("/timesheets/<timesheet_id>/reject", methods=["POST"])
@login_required
@can_approve
def reject_timesheet(timesheet_id):
    """
    Mark a timesheet as needing approval (missing attachment).
    
    REQ-041: Support users can only reject trainee timesheets.

    Request body:
        reason: Optional rejection reason (also sets admin_notes)

    Returns:
        dict: Updated timesheet
    """
    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    if timesheet.status != TimesheetStatus.SUBMITTED:
        return {"error": "Only submitted timesheets can be rejected"}, 400

    if get_confirmed_pay_period(timesheet.week_start):
        return {"error": "Pay period has been confirmed and is locked"}, 400

    timesheet.status = TimesheetStatus.NEEDS_APPROVAL

    # Add note if reason provided - also set admin_notes field
    data = request.get_json() or {}
    reason = data.get("reason", "").strip()
    if reason:
        approver_id = session["user"]["id"]
        # Set the admin_notes field
        timesheet.admin_notes = reason
        # Also create a Note record for history
        note = Note(
            timesheet_id=timesheet_id,
            author_id=approver_id,
            content=f"Needs approval: {reason}",
        )
        db.session.add(note)

    db.session.commit()

    # Send SMS notification
    from ..services.notification import NotificationService

    NotificationService.notify_needs_attention(timesheet, reason)

    return timesheet.to_dict()


@admin_bp.route("/timesheets/<timesheet_id>/admin-notes", methods=["PUT"])
@login_required
@can_approve
def update_admin_notes(timesheet_id):
    """
    Update admin notes on a timesheet.
    
    REQ-041: Support users can only update notes on trainee timesheets.

    Request body:
        admin_notes: The admin notes text

    Returns:
        dict: Updated timesheet
    """
    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404

    # Cannot edit drafts
    if timesheet.status == TimesheetStatus.NEW:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    data = request.get_json() or {}
    timesheet.admin_notes = data.get("admin_notes", "").strip() or None
    db.session.commit()

    return timesheet.to_dict()


@admin_bp.route("/timesheets/<timesheet_id>/unapprove", methods=["POST"])
@login_required
@can_approve
def unapprove_timesheet(timesheet_id):
    """
    Revert an approved timesheet back to submitted status.
    
    REQ-041: Support users can only unapprove trainee timesheets.

    Returns:
        dict: Updated timesheet
    """
    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    if timesheet.status != TimesheetStatus.APPROVED:
        return {"error": "Only approved timesheets can be unapproved"}, 400

    if get_confirmed_pay_period(timesheet.week_start):
        return {"error": "Pay period has been confirmed and is locked"}, 400

    timesheet.status = TimesheetStatus.SUBMITTED
    timesheet.approved_at = None
    timesheet.approved_by = None
    db.session.commit()

    return timesheet.to_dict()


@admin_bp.route("/pay-periods/status", methods=["GET"])
@login_required
@admin_required
def get_pay_period_status():
    """
    Get confirmation status for a pay period.

    Query params:
        start_date: Pay period start (YYYY-MM-DD)
        end_date: Pay period end (YYYY-MM-DD)
    """
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    if not start_date or not end_date:
        return {"error": "start_date and end_date are required"}, 400

    try:
        start = datetime.fromisoformat(start_date).date()
        end = datetime.fromisoformat(end_date).date()
    except ValueError:
        return {"error": "Invalid date format"}, 400

    period = PayPeriod.query.filter_by(start_date=start).first()
    if period and period.end_date != end:
        return {"error": "Pay period dates do not match existing record"}, 400

    return {
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "confirmed": period is not None,
        "pay_period": period.to_dict() if period else None,
    }


@admin_bp.route("/pay-periods/confirm", methods=["POST"])
@login_required
@admin_required
def confirm_pay_period():
    """
    Confirm and lock a pay period (REQ-006).

    Request body:
        start_date: Pay period start (YYYY-MM-DD)
        end_date: Pay period end (YYYY-MM-DD)
    """
    data = request.get_json() or {}
    start_date = data.get("start_date")
    end_date = data.get("end_date")

    if not start_date or not end_date:
        return {"error": "start_date and end_date are required"}, 400

    try:
        start = datetime.fromisoformat(start_date).date()
        end = datetime.fromisoformat(end_date).date()
    except ValueError:
        return {"error": "Invalid date format"}, 400

    if start.weekday() != 0 or (end - start).days != 13:
        return {"error": "Pay period must start on Monday and span 14 days"}, 400

    existing = PayPeriod.query.filter_by(start_date=start).first()
    if existing:
        return {"error": "Pay period already confirmed"}, 400

    timesheets = Timesheet.query.filter(
        Timesheet.week_start >= start,
        Timesheet.week_start <= end,
    ).all()

    not_approved = [ts for ts in timesheets if ts.status != TimesheetStatus.APPROVED]
    if not_approved:
        status_counts = {}
        for ts in not_approved:
            status_counts[ts.status] = status_counts.get(ts.status, 0) + 1
        return {
            "error": "All timesheets must be approved before confirmation",
            "details": {
                "pending_count": len(not_approved),
                "status_counts": status_counts,
            },
        }, 400

    pay_period = PayPeriod(
        start_date=start,
        end_date=end,
        confirmed_by=session["user"]["id"],
    )
    db.session.add(pay_period)
    db.session.commit()

    return pay_period.to_dict(), 201


def _apply_role_scope(query):
    current_role = session.get("user", {}).get("role", "staff")
    is_support_only = current_role == "support"

    query = query.join(User, Timesheet.user_id == User.id)

    if is_support_only:
        query = query.filter(User.role == UserRole.TRAINEE)

    return query


def _build_export_query():
    query = Timesheet.query.filter(Timesheet.status != TimesheetStatus.NEW)
    query = _apply_role_scope(query)

    status = request.args.get("status")
    if status and status in [
        TimesheetStatus.SUBMITTED,
        TimesheetStatus.APPROVED,
        TimesheetStatus.NEEDS_APPROVAL,
    ]:
        query = query.filter(Timesheet.status == status)

    user_id = request.args.get("user_id")
    if user_id:
        query = query.filter(Timesheet.user_id == user_id)

    week_start = request.args.get("week_start")
    if week_start:
        query = query.filter(
            Timesheet.week_start == datetime.fromisoformat(week_start).date()
        )

    pay_period_start = request.args.get("pay_period_start")
    pay_period_end = request.args.get("pay_period_end")
    if pay_period_start and pay_period_end:
        start_date = datetime.fromisoformat(pay_period_start).date()
        end_date = datetime.fromisoformat(pay_period_end).date()
        query = query.filter(Timesheet.week_start >= start_date).filter(
            Timesheet.week_start <= end_date
        )

    hour_type = request.args.get("hour_type")
    if hour_type:
        from ..models import TimesheetEntry

        if hour_type == "has_field":
            query = query.filter(
                Timesheet.id.in_(
                    db.session.query(TimesheetEntry.timesheet_id)
                    .filter(TimesheetEntry.hour_type == "Field")
                    .distinct()
                )
            )
        else:
            query = query.filter(
                Timesheet.id.in_(
                    db.session.query(TimesheetEntry.timesheet_id)
                    .filter(TimesheetEntry.hour_type == hour_type)
                    .distinct()
                )
            )

    return query.order_by(Timesheet.week_start.desc())


def _parse_export_format():
    export_format = (request.args.get("format") or "csv").lower()
    if export_format not in ("csv", "xlsx", "pdf"):
        return None
    return export_format


def _summary_headers():
    return [
        "Employee",
        "Email",
        "Week Start",
        "Status",
        "Total Hours",
        "Payable Hours",
        "Billable Hours",
        "Unpaid Hours",
        "Traveled",
        "Expenses",
        "Reimbursement",
        "Attachments",
        "Created At",
    ]


def _summary_row(timesheet):
    totals = timesheet.calculate_totals()
    reimbursement = ""
    if timesheet.reimbursement_needed:
        amount = float(timesheet.reimbursement_amount or 0)
        reimbursement = f"${amount:.2f}"

    attachments_count = timesheet.attachments.count()

    return [
        timesheet.user.display_name if timesheet.user else "Unknown",
        timesheet.user.email if timesheet.user else "",
        timesheet.week_start.isoformat(),
        timesheet.status,
        float(totals["total"]),
        float(totals["payable"]),
        float(totals["billable"]),
        float(totals["unpaid"]),
        "Yes" if timesheet.traveled else "No",
        "Yes" if timesheet.has_expenses else "No",
        reimbursement,
        attachments_count,
        timesheet.created_at.date().isoformat(),
    ]


def _totals_row(timesheets):
    total_hours = 0.0
    payable_hours = 0.0
    billable_hours = 0.0
    unpaid_hours = 0.0
    attachments_total = 0

    for ts in timesheets:
        totals = ts.calculate_totals()
        total_hours += float(totals["total"])
        payable_hours += float(totals["payable"])
        billable_hours += float(totals["billable"])
        unpaid_hours += float(totals["unpaid"])
        attachments_total += ts.attachments.count()

    return [
        "TOTALS",
        "",
        "",
        "",
        total_hours,
        payable_hours,
        billable_hours,
        unpaid_hours,
        "",
        "",
        "",
        attachments_total,
        "",
    ]


def _send_csv(headers, rows, filename, totals_row=None, title=None):
    output = StringIO()
    writer = csv.writer(output)
    if title:
        writer.writerow([title])
        writer.writerow([])
    writer.writerow(headers)
    writer.writerows(rows)
    if totals_row:
        writer.writerow([])
        writer.writerow(totals_row)

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _send_excel(headers, rows, filename, totals_row=None, title=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return {"error": "Excel export requires openpyxl"}, 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheets"

    if title:
        ws.append([title])
        ws.append([])

    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    if totals_row:
        ws.append([])
        ws.append(totals_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _send_pdf(headers, rows, filename, title, totals_row=None, extra_tables=None):
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return {"error": "PDF export requires reportlab"}, 500

    buffer = BytesIO()
    # Use smaller margins to fit all columns
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        leftMargin=0.3*inch,
        rightMargin=0.3*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    table_data = [headers] + rows
    if totals_row:
        table_data.append(totals_row)

    # Calculate column widths based on content
    # Landscape letter is 11" wide, minus margins = ~10.4" usable
    num_cols = len(headers)
    available_width = 10.4 * inch
    col_width = available_width / num_cols

    table = Table(table_data, repeatRows=1, colWidths=[col_width] * num_cols)
    
    # Build table styles
    table_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),  # Smaller font for better fit
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    
    # Style the totals row distinctly (last row if totals_row was added)
    if totals_row:
        last_row = len(table_data) - 1
        table_styles.extend([
            ("BACKGROUND", (0, last_row), (-1, last_row), colors.HexColor("#e5e7eb")),
            ("FONTNAME", (0, last_row), (-1, last_row), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, last_row), (-1, last_row), colors.HexColor("#1f2937")),
        ])
    
    table.setStyle(TableStyle(table_styles))
    elements.append(table)

    if extra_tables:
        elements.append(Spacer(1, 18))
        for extra in extra_tables:
            elements.append(Paragraph(extra["title"], styles["Heading3"]))
            elements.append(Spacer(1, 6))
            extra_table = Table(extra["data"], repeatRows=1)
            extra_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ]
                )
            )
            elements.append(extra_table)
            elements.append(Spacer(1, 12))

    doc.build(elements)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@admin_bp.route("/exports/timesheets", methods=["GET"])
@login_required
@can_approve
def export_timesheets():
    """
    Export filtered timesheets in CSV/XLSX/PDF format (REQ-019).
    """
    export_format = _parse_export_format()
    if not export_format:
        return {"error": "Invalid export format"}, 400

    try:
        query = _build_export_query()
    except ValueError:
        return {"error": "Invalid date format"}, 400

    timesheets = query.all()
    rows = [_summary_row(ts) for ts in timesheets]
    totals_row = _totals_row(timesheets) if timesheets else None

    today = datetime.utcnow().date().isoformat()
    filename = f"timesheets_export_{today}.{export_format}"
    title = "Timesheet Export"

    if export_format == "csv":
        return _send_csv(_summary_headers(), rows, filename, totals_row, title=title)
    if export_format == "xlsx":
        return _send_excel(_summary_headers(), rows, filename, totals_row, title=title)
    return _send_pdf(_summary_headers(), rows, filename, title, totals_row)


@admin_bp.route("/exports/timesheets/<timesheet_id>", methods=["GET"])
@login_required
@can_approve
def export_timesheet_detail(timesheet_id):
    """
    Export a single timesheet detail (REQ-019).
    """
    export_format = _parse_export_format()
    if not export_format:
        return {"error": "Invalid export format"}, 400

    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()
    if not timesheet or timesheet.status == TimesheetStatus.NEW:
        return {"error": "Timesheet not found"}, 404

    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    summary_rows = [
        ["Employee", timesheet.user.display_name if timesheet.user else "Unknown"],
        ["Email", timesheet.user.email if timesheet.user else ""],
        ["Week Start", timesheet.week_start.isoformat()],
        ["Status", timesheet.status],
        ["Total Hours", float(timesheet.calculate_totals()["total"])],
        ["Payable Hours", float(timesheet.calculate_totals()["payable"])],
        ["Billable Hours", float(timesheet.calculate_totals()["billable"])],
        ["Unpaid Hours", float(timesheet.calculate_totals()["unpaid"])],
        ["Traveled", "Yes" if timesheet.traveled else "No"],
        ["Has Expenses", "Yes" if timesheet.has_expenses else "No"],
        [
            "Reimbursement",
            f"{timesheet.reimbursement_type or ''} ${float(timesheet.reimbursement_amount or 0):.2f}"
            if timesheet.reimbursement_needed
            else "No",
        ],
        ["Attachments", timesheet.attachments.count()],
    ]

    entry_rows = [
        [entry.entry_date.isoformat(), entry.hour_type, float(entry.hours)]
        for entry in timesheet.entries.order_by(TimesheetEntry.entry_date).all()
    ]

    filename = f"timesheet_{timesheet.week_start.isoformat()}_{timesheet_id}.{export_format}"
    title = f"Timesheet Detail - {timesheet.week_start.isoformat()}"

    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([title])
        writer.writerow([])
        writer.writerow(["Field", "Value"])
        writer.writerows(summary_rows)
        writer.writerow([])
        writer.writerow(["Entries"])
        writer.writerow(["Date", "Hour Type", "Hours"])
        writer.writerows(entry_rows)
        response = Response(output.getvalue(), mimetype="text/csv")
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            return {"error": "Excel export requires openpyxl"}, 500

        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        summary_sheet.append([title])
        summary_sheet.append([])
        summary_sheet.append(["Field", "Value"])
        for cell in summary_sheet[summary_sheet.max_row]:
            cell.font = Font(bold=True)
        for row in summary_rows:
            summary_sheet.append(row)

        entries_sheet = wb.create_sheet(title="Entries")
        entries_sheet.append(["Date", "Hour Type", "Hours"])
        for cell in entries_sheet[1]:
            cell.font = Font(bold=True)
        for row in entry_rows:
            entries_sheet.append(row)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    extra_tables = [
        {
            "title": "Entries",
            "data": [["Date", "Hour Type", "Hours"]] + entry_rows,
        }
    ]
    return _send_pdf(["Field", "Value"], summary_rows, filename, title, extra_tables=extra_tables)


@admin_bp.route("/exports/pay-period", methods=["GET"])
@login_required
@admin_required
def export_pay_period():
    """
    Export a pay period summary (REQ-019).
    """
    export_format = _parse_export_format()
    if not export_format:
        return {"error": "Invalid export format"}, 400

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    if not start_date or not end_date:
        return {"error": "start_date and end_date are required"}, 400

    try:
        start = datetime.fromisoformat(start_date).date()
        end = datetime.fromisoformat(end_date).date()
    except ValueError:
        return {"error": "Invalid date format"}, 400

    query = Timesheet.query.filter(Timesheet.status != TimesheetStatus.NEW)
    query = _apply_role_scope(query)
    query = query.filter(Timesheet.week_start >= start).filter(
        Timesheet.week_start <= end
    )

    timesheets = query.order_by(Timesheet.week_start).all()
    rows = [_summary_row(ts) for ts in timesheets]
    totals_row = _totals_row(timesheets) if timesheets else None

    title = f"Pay Period Summary: {start.isoformat()} to {end.isoformat()}"
    filename = f"pay_period_{start.isoformat()}_{end.isoformat()}.{export_format}"

    if export_format == "csv":
        return _send_csv(_summary_headers(), rows, filename, totals_row, title=title)
    if export_format == "xlsx":
        return _send_excel(_summary_headers(), rows, filename, totals_row, title=title)
    return _send_pdf(_summary_headers(), rows, filename, title, totals_row)


@admin_bp.route(
    "/timesheets/<timesheet_id>/attachments/<attachment_id>", methods=["GET"]
)
@login_required
@can_approve
def download_attachment(timesheet_id, attachment_id):
    """
    Download an attachment for review.
    
    REQ-041: Support users can only download attachments from trainee timesheets.

    Returns:
        file: The attachment file
    """
    import os
    from ..models import Attachment

    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404

    # Cannot view draft attachments
    if timesheet.status == TimesheetStatus.NEW:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    attachment = Attachment.query.filter_by(
        id=attachment_id, timesheet_id=timesheet_id
    ).first()

    if not attachment:
        return {"error": "Attachment not found"}, 404

    filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], attachment.filename)

    if not os.path.exists(filepath):
        return {"error": "File not found"}, 404

    return send_file(
        filepath,
        mimetype=attachment.mime_type,
        as_attachment=True,
        download_name=attachment.original_filename,
    )


@admin_bp.route("/attachments/<attachment_id>/sharepoint/retry", methods=["POST"])
@login_required
@admin_required
def retry_sharepoint_sync(attachment_id):
    """
    Retry SharePoint sync for a specific attachment (REQ-010).
    """
    from ..jobs import enqueue_sharepoint_sync

    attachment = Attachment.query.filter_by(id=attachment_id).first()
    if not attachment:
        return {"error": "Attachment not found"}, 404

    if not current_app.config.get("SHAREPOINT_SYNC_ENABLED", False):
        return {"error": "SharePoint sync is disabled"}, 400

    attachment.sharepoint_sync_status = Attachment.SharePointSyncStatus.PENDING
    attachment.sharepoint_last_error = None
    db.session.commit()

    enqueue_sharepoint_sync(attachment.id)
    return {"status": "queued"}


@admin_bp.route("/timesheets/<timesheet_id>/notes", methods=["POST"])
@login_required
@can_approve
def add_note(timesheet_id):
    """
    Add a note to a timesheet.
    
    REQ-041: Support users can only add notes to trainee timesheets.

    Request body:
        content: Note text

    Returns:
        dict: Created note
    """
    author_id = session["user"]["id"]

    timesheet = Timesheet.query.filter_by(id=timesheet_id).first()

    if not timesheet:
        return {"error": "Timesheet not found"}, 404

    # Cannot add notes to drafts
    if timesheet.status == TimesheetStatus.NEW:
        return {"error": "Timesheet not found"}, 404
    
    # REQ-041: Check if Support user can access this timesheet
    can_access, error = _can_access_timesheet(timesheet)
    if not can_access:
        return error

    data = request.get_json() or {}
    content = data.get("content", "").strip()

    if not content:
        return {"error": "Note content required"}, 400

    note = Note(
        timesheet_id=timesheet_id,
        author_id=author_id,
        content=content,
    )
    db.session.add(note)
    db.session.commit()

    return note.to_dict(), 201


@admin_bp.route("/users", methods=["GET"])
@login_required
@admin_required
def list_users():
    """
    List all users (admin only).

    Returns:
        dict: List of users
    """
    users = User.query.order_by(User.display_name).all()

    return {
        "users": [u.to_dict() for u in users],
    }
